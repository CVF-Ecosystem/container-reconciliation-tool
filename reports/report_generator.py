# File: report_generator.py (V5.1 - Enhanced Movement Summary & Checks)
import pandas as pd
from pathlib import Path
import logging
from typing import Dict, Any, Optional
from config import Col

try:
    from reports.movement_summary import generate_full_movement_report
except ImportError:
    generate_full_movement_report = None

def _auto_format_excel(writer: pd.ExcelWriter, df: pd.DataFrame, sheet_name: str) -> None:
    """
    Tự động định dạng file Excel cho dễ đọc.
    V5.6: Compatible with both xlsxwriter and openpyxl engines.
    """
    try:
        workbook = writer.book
        
        # Check if using xlsxwriter (has add_format method)
        if hasattr(workbook, 'add_format'):
            # xlsxwriter engine
            worksheet = writer.sheets[sheet_name]
            
            # Define formats
            header_format = workbook.add_format({
                'bold': True,
                'bg_color': '#4472C4',
                'font_color': 'white',
                'border': 1,
                'align': 'center',
                'valign': 'vcenter'
            })
            
            # Apply header format
            for col_num, value in enumerate(df.columns.values):
                worksheet.write(0, col_num, value, header_format)
            
            # Auto-adjust column widths
            for i, col in enumerate(df.columns):
                max_len = max(
                    df[col].astype(str).apply(len).max(),
                    len(str(col))
                ) + 2
                worksheet.set_column(i, i, min(max_len, 50))
            
            # Freeze top row
            worksheet.freeze_panes(1, 0)
            
        else:
            # openpyxl engine - use different formatting approach
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            worksheet = writer.sheets[sheet_name]
            
            # Header formatting
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_align = Alignment(horizontal='center', vertical='center')
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Apply header format
            for col_num, value in enumerate(df.columns.values, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border
            
            # Auto-adjust column widths
            for i, col in enumerate(df.columns, 1):
                max_len = max(
                    df[col].astype(str).apply(len).max(),
                    len(str(col))
                ) + 2
                worksheet.column_dimensions[get_column_letter(i)].width = min(max_len, 50)
            
            # Freeze top row
            worksheet.freeze_panes = 'A2'
        
    except Exception as e:
        logging.warning(f"Could not apply Excel formatting: {e}")

def _add_total_row(df: pd.DataFrame) -> pd.DataFrame:
    """Thêm dòng TOTAL vào cuối DataFrame cho các cột số."""
    if df.empty:
        return df
    
    # Tìm các cột số để tính tổng
    numeric_cols = df.select_dtypes(include=['int64', 'float64', 'int32', 'float32']).columns.tolist()
    
    if not numeric_cols:
        return df
    
    # Tạo dòng tổng
    total_row = {}
    for col in df.columns:
        if col in numeric_cols:
            total_row[col] = df[col].sum()
        elif col == df.columns[0]:
            total_row[col] = 'TỔNG CỘNG'
        else:
            total_row[col] = ''
    
    # Thêm dòng tổng vào DataFrame
    df_with_total = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)
    return df_with_total

def _write_sheet(writer: pd.ExcelWriter, df: pd.DataFrame, sheet_name: str, include_index: bool = False, add_total: bool = True) -> None:
    """Helper để ghi 1 sheet với format và dòng tổng."""
    if df is None or df.empty:
        return
    sheet_name_clean = sheet_name.replace(" ", "_")[:31]
    
    # Thêm dòng tổng nếu cần
    if add_total:
        df_to_write = _add_total_row(df.copy())
    else:
        df_to_write = df
    
    df_to_write.to_excel(writer, sheet_name=sheet_name_clean, index=include_index)
    _auto_format_excel(writer, df_to_write if not include_index else df_to_write.reset_index(), sheet_name_clean)

def _create_phuong_an_breakdown(raw_data: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    """
    V4.5.1: Tạo bảng thống kê chi tiết theo nguồn và Phương án.
    """
    breakdown_data = []
    
    # Mapping source key to readable name - Việt hóa hoàn toàn
    source_names = {
        'ton_cu': 'TỒN ĐẦU KỲ (Baseline)',
        'ton_moi': 'TỒN HIỆN TẠI',
        'gate_in': 'CỔNG VÀO (Gate In)',
        'gate_out': 'CỔNG RA (Gate Out)',
        'nhap_tau': 'NHẬP TÀU (Discharge)',
        'xuat_tau': 'XUẤT TÀU (Loading)',
        'nhap_shifting': 'SHIFTING NHẬP (Tàu→Bãi)',
        'xuat_shifting': 'SHIFTING XUẤT (Bãi→Tàu)',
    }
    
    for source_key, source_name in source_names.items():
        df = raw_data.get(source_key, pd.DataFrame())
        if df.empty:
            continue
            
        total_count = len(df)
        
        # Thống kê theo Phương án
        if Col.PHUONG_AN in df.columns:
            phuong_an_counts = df[Col.PHUONG_AN].value_counts()
            for phuong_an, count in phuong_an_counts.items():
                if pd.isna(phuong_an) or str(phuong_an).strip() == '':
                    phuong_an = '(Không có phương án)'
                breakdown_data.append({
                    'Nguồn': source_name,
                    'Phương án': str(phuong_an),
                    'Số lượng': count,
                    'Hướng': 'VÀO' if source_key in ['ton_cu', 'gate_in', 'nhap_tau', 'nhap_shifting', 'ton_moi'] else 'RA'
                })
        else:
            # Không có cột Phương án
            breakdown_data.append({
                'Nguồn': source_name,
                'Phương án': '(Tổng)',
                'Số lượng': total_count,
                'Hướng': '-'
            })
    
    if not breakdown_data:
        return pd.DataFrame()
    
    df_breakdown = pd.DataFrame(breakdown_data)
    # Sắp xếp theo nguồn, rồi theo số lượng giảm dần
    df_breakdown = df_breakdown.sort_values(['Nguồn', 'Số lượng'], ascending=[True, False])
    
    return df_breakdown

def _create_summary_by_source(raw_data: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    """
    Tạo bảng tổng hợp theo nguồn (không chi tiết phương án).
    """
    summary_data = []
    
    # Group 1: TỒN BÃI
    summary_data.append({'Nhóm': 'TỒN BÃI', 'Loại': 'Tồn đầu kỳ (Baseline)', 'Số lượng': len(raw_data.get('ton_cu', pd.DataFrame()))})
    summary_data.append({'Nhóm': 'TỒN BÃI', 'Loại': 'Tồn hiện tại', 'Số lượng': len(raw_data.get('ton_moi', pd.DataFrame()))})
    
    # Group 2: CỔNG (GATE)
    summary_data.append({'Nhóm': 'CỔNG', 'Loại': 'Cổng vào', 'Số lượng': len(raw_data.get('gate_in', pd.DataFrame()))})
    summary_data.append({'Nhóm': 'CỔNG', 'Loại': 'Cổng ra', 'Số lượng': len(raw_data.get('gate_out', pd.DataFrame()))})
    
    # Group 3: TÀU
    summary_data.append({'Nhóm': 'TÀU', 'Loại': 'Nhập tàu (Discharge)', 'Số lượng': len(raw_data.get('nhap_tau', pd.DataFrame()))})
    summary_data.append({'Nhóm': 'TÀU', 'Loại': 'Xuất tàu (Loading)', 'Số lượng': len(raw_data.get('xuat_tau', pd.DataFrame()))})
    
    # Group 4: SHIFTING
    summary_data.append({'Nhóm': 'SHIFTING', 'Loại': 'Shifting nhập (Tàu→Bãi)', 'Số lượng': len(raw_data.get('nhap_shifting', pd.DataFrame()))})
    summary_data.append({'Nhóm': 'SHIFTING', 'Loại': 'Shifting xuất (Bãi→Tàu)', 'Số lượng': len(raw_data.get('xuat_shifting', pd.DataFrame()))})
    
    return pd.DataFrame(summary_data)

def _create_inventory_change_summary(inventory_change_results: Dict) -> pd.DataFrame:
    """
    Tổng hợp biến động TỒN CŨ vs TỒN MỚI theo nguồn gốc.
    """
    summary_data = []
    
    df_moi_vao = inventory_change_results.get('moi_vao_bai', pd.DataFrame())
    df_da_roi = inventory_change_results.get('da_roi_bai', pd.DataFrame())
    df_van_ton = inventory_change_results.get('van_con_ton', pd.DataFrame())
    
    # Tổng quan
    summary_data.append({'Hạng mục': 'Container MỚI VÀO bãi', 'Số lượng': len(df_moi_vao)})
    summary_data.append({'Hạng mục': 'Container ĐÃ RỜI bãi', 'Số lượng': len(df_da_roi)})
    summary_data.append({'Hạng mục': 'Container VẪN TỒN', 'Số lượng': len(df_van_ton)})
    
    # Chi tiết nguồn gốc MỚI VÀO
    if not df_moi_vao.empty and 'NguonGoc' in df_moi_vao.columns:
        summary_data.append({'Hạng mục': '--- Chi tiết MỚI VÀO ---', 'Số lượng': ''})
        for source, count in df_moi_vao['NguonGoc'].value_counts().items():
            # Nếu là "Không rõ", liệt kê chi tiết các container
            if source == 'Không rõ' and count > 0:
                containers = df_moi_vao[df_moi_vao['NguonGoc'] == 'Không rõ'][Col.CONTAINER].tolist()
                cont_list = ', '.join(containers[:10])  # Hiện tối đa 10
                if len(containers) > 10:
                    cont_list += f'... (+{len(containers)-10})'
                summary_data.append({'Hạng mục': f'  + {source}', 'Số lượng': f'{count} [{cont_list}]'})
            else:
                summary_data.append({'Hạng mục': f'  + {source}', 'Số lượng': count})
    
    # Chi tiết nguồn gốc ĐÃ RỜI
    if not df_da_roi.empty and 'NguonGoc' in df_da_roi.columns:
        summary_data.append({'Hạng mục': '--- Chi tiết ĐÃ RỜI ---', 'Số lượng': ''})
        for source, count in df_da_roi['NguonGoc'].value_counts().items():
            # Nếu là "Không rõ", liệt kê chi tiết các container
            if source == 'Không rõ' and count > 0:
                containers = df_da_roi[df_da_roi['NguonGoc'] == 'Không rõ'][Col.CONTAINER].tolist()
                cont_list = ', '.join(containers[:10])  # Hiện tối đa 10
                if len(containers) > 10:
                    cont_list += f'... (+{len(containers)-10})'
                summary_data.append({'Hạng mục': f'  - {source}', 'Số lượng': f'{count} [{cont_list}]'})
            else:
                summary_data.append({'Hạng mục': f'  - {source}', 'Số lượng': count})
    
    return pd.DataFrame(summary_data)

def create_reports(all_results: Dict[str, Any]) -> None:
    """
    V4.5.1: Tạo báo cáo với SUMMARY chi tiết theo Phương án.
    """
    report_folder = all_results["report_folder"]
    main_results = all_results["main_results"]
    operator_analysis_result = all_results["operator_analysis_result"]
    delta_analysis_result = all_results["delta_analysis_result"]
    summary_df = all_results["summary_df"]
    quality_warnings = all_results["quality_warnings"]
    inventory_change_results = all_results["inventory_change_results"]
    raw_data = main_results.get('raw_data', {})
    
    logging.info(f"--- GIAI ĐOẠN 3: TẠO BÁO CÁO TẠI {report_folder} ---")
    
    # ========================================
    # FILE 1: SUMMARY.xlsx - Tổng hợp tất cả
    # ========================================
    summary_path = report_folder / "1. SUMMARY.xlsx"
    try:
        with pd.ExcelWriter(summary_path, engine='openpyxl') as writer:
            # Sheet 1: Tổng hợp chính
            _write_sheet(writer, summary_df, "1_Tong_Hop_Chinh")
            
            # Sheet 2: Thống kê theo nguồn dữ liệu
            df_source_summary = _create_summary_by_source(raw_data)
            _write_sheet(writer, df_source_summary, "2_Theo_Nguon")
            
            # Sheet 3: Chi tiết theo Phương án
            df_phuong_an = _create_phuong_an_breakdown(raw_data)
            _write_sheet(writer, df_phuong_an, "3_Chi_Tiet_Phuong_An")
            
            # Sheet 4: Biến động tồn bãi
            df_inventory_summary = _create_inventory_change_summary(inventory_change_results)
            _write_sheet(writer, df_inventory_summary, "4_Bien_Dong_Ton_Bai")
            
            # Sheet 5: Delta Analysis (so sánh lần trước)
            if delta_analysis_result is not None and not delta_analysis_result.empty:
                _write_sheet(writer, delta_analysis_result, "5_So_Sanh_Lan_Truoc")
            
            # Sheet 6: Cảnh báo chất lượng
            df_quality = pd.DataFrame(quality_warnings, columns=['Vấn đề Chất lượng Dữ liệu'])
            df_future = main_results.get('future_moves_report', pd.DataFrame())
            df_suspicious = main_results.get('suspicious_dates', pd.DataFrame())
            if not df_quality.empty:
                _write_sheet(writer, df_quality, "6_Canh_Bao_Chat_Luong")
            
            # Sheet 7: Chi tiết Ngày tương lai (QUAN TRỌNG để kiểm tra)
            if not df_future.empty:
                # Chỉ giữ các cột cần thiết để tra cứu
                cols_to_keep = [Col.CONTAINER, Col.SOURCE_FILE, Col.SOURCE_KEY, Col.TRANSACTION_TIME, 
                               Col.PHUONG_AN, Col.OPERATOR, 'GhiChu_SuaLoi']
                cols_available = [c for c in cols_to_keep if c in df_future.columns]
                df_future_export = df_future[cols_available].copy()
                df_future_export = df_future_export.sort_values(Col.TRANSACTION_TIME, ascending=False)
                _write_sheet(writer, df_future_export, "7_Ngay_Tuong_Lai_CHI_TIET", add_total=False)
                logging.warning(f"[SUMMARY] Đã xuất {len(df_future)} container có ngày tương lai để kiểm tra!")
            
            # Sheet 8: Chi tiết Ngày đáng ngờ (tháng/ngày bị đảo?)
            if not df_suspicious.empty:
                cols_to_keep = [Col.CONTAINER, Col.SOURCE_FILE, Col.SOURCE_KEY, Col.TRANSACTION_TIME, 
                               Col.PHUONG_AN, Col.OPERATOR]
                cols_available = [c for c in cols_to_keep if c in df_suspicious.columns]
                df_suspicious_export = df_suspicious[cols_available].copy()
                _write_sheet(writer, df_suspicious_export, "8_Ngay_Dang_Ngo_CHI_TIET", add_total=False)
                
    except Exception as e:
        logging.error(f"Lỗi khi tạo SUMMARY: {e}")
    
    # ========================================
    # FILE 2: TON_BAI_CHUAN.xlsx - Container khớp
    # ========================================
    ton_chuan_path = report_folder / "2. TON_BAI_CHUAN.xlsx"
    try:
        with pd.ExcelWriter(ton_chuan_path, engine='xlsxwriter') as writer:
            _write_sheet(writer, main_results.get('ton_chuan'), "Ton_Chuan")
    except Exception as e:
        logging.error(f"Lỗi khi tạo TON_BAI_CHUAN: {e}")
    
    # ========================================
    # FILE 3: CHENH_LECH.xlsx - Container (+) và (-)
    # ========================================
    chenh_lech_path = report_folder / "3. CHENH_LECH.xlsx"
    try:
        with pd.ExcelWriter(chenh_lech_path, engine='xlsxwriter') as writer:
            _write_sheet(writer, main_results.get('chenh_lech_am'), "Co_Lenh_Chua_Ve")
            _write_sheet(writer, main_results.get('chenh_lech_duong'), "Ton_Chua_Co_Lenh")
            _write_sheet(writer, main_results.get('bien_dong_fe'), "Bien_Dong_FE_CFS")
            _write_sheet(writer, main_results.get('sai_thong_tin'), "Sai_Thong_Tin")
            _write_sheet(writer, main_results.get('xuat_tau_van_ton'), "Xuat_Tau_Van_Ton_LOI")
    except Exception as e:
        logging.error(f"Lỗi khi tạo CHENH_LECH: {e}")
    
    # ========================================
    # FILE 4: DAO_CHUYEN_NOI_BAI.xlsx - Position Shift
    # ========================================
    dao_chuyen_path = report_folder / "4. DAO_CHUYEN_NOI_BAI.xlsx"
    try:
        with pd.ExcelWriter(dao_chuyen_path, engine='xlsxwriter') as writer:
            _write_sheet(writer, main_results.get('dao_chuyen_noi_bai'), "Dao_Chuyen_Vi_Tri")
            _write_sheet(writer, main_results.get('pending_shifting'), "Restow_Dang_Cho")
    except Exception as e:
        logging.error(f"Lỗi khi tạo DAO_CHUYEN_NOI_BAI: {e}")
    
    # ========================================
    # FILE 5: BIEN_DONG_CHI_TIET.xlsx - VÀO/RA theo nguồn
    # ========================================
    bien_dong_path = report_folder / "5. BIEN_DONG_CHI_TIET.xlsx"
    try:
        with pd.ExcelWriter(bien_dong_path, engine='xlsxwriter') as writer:
            df_moi_vao = inventory_change_results.get('moi_vao_bai', pd.DataFrame())
            df_da_roi = inventory_change_results.get('da_roi_bai', pd.DataFrame())
            
            # Summary by Phương án for VÀO
            if not df_moi_vao.empty and Col.PHUONG_AN in df_moi_vao.columns:
                vao_by_pa = df_moi_vao.groupby(Col.PHUONG_AN).size().reset_index(name='Số lượng')
                vao_by_pa.columns = ['Phương án', 'Số lượng']
                vao_by_pa = vao_by_pa.sort_values('Số lượng', ascending=False)
                _write_sheet(writer, vao_by_pa, "Vao_Theo_Phuong_An")
            
            # Summary by Phương án for RA  
            if not df_da_roi.empty and Col.PHUONG_AN in df_da_roi.columns:
                ra_by_pa = df_da_roi.groupby(Col.PHUONG_AN).size().reset_index(name='Số lượng')
                ra_by_pa.columns = ['Phương án', 'Số lượng']
                ra_by_pa = ra_by_pa.sort_values('Số lượng', ascending=False)
                _write_sheet(writer, ra_by_pa, "Ra_Theo_Phuong_An")
            
            _write_sheet(writer, df_moi_vao, "Chi_Tiet_Moi_Vao")
            _write_sheet(writer, df_da_roi, "Chi_Tiet_Da_Roi")
            
            # Sheet riêng cho container "Không rõ" nguồn gốc
            if not df_moi_vao.empty and 'NguonGoc' in df_moi_vao.columns:
                df_khong_ro_vao = df_moi_vao[df_moi_vao['NguonGoc'] == 'Không rõ']
                if not df_khong_ro_vao.empty:
                    _write_sheet(writer, df_khong_ro_vao, "Khong_Ro_Vao_KIEM_TRA")
                    logging.warning(f"[BIEN_DONG] {len(df_khong_ro_vao)} container VÀO không rõ nguồn gốc!")
            
            if not df_da_roi.empty and 'NguonGoc' in df_da_roi.columns:
                df_khong_ro_ra = df_da_roi[df_da_roi['NguonGoc'] == 'Không rõ']
                if not df_khong_ro_ra.empty:
                    _write_sheet(writer, df_khong_ro_ra, "Khong_Ro_Ra_KIEM_TRA")
                    logging.warning(f"[BIEN_DONG] {len(df_khong_ro_ra)} container RA không rõ nguồn gốc!")
                    
    except Exception as e:
        logging.error(f"Lỗi khi tạo BIEN_DONG_CHI_TIET: {e}")
    
    # ========================================
    # FILE 6: TON_THEO_HANG - Tách file riêng cho từng hãng (V5.1.4)
    # ========================================
    # Tạo thư mục con cho các file theo hãng
    hang_folder = report_folder / "6_Ton_Theo_Hang"
    hang_folder.mkdir(exist_ok=True)
    
    # Import time slot filter
    try:
        from utils.time_slot_filter import get_operator_time_config, get_shifts_for_operator, get_shift_display_name
        has_time_filter = True
    except ImportError:
        has_time_filter = False
        logging.warning("[TON_THEO_HANG] Không có time_slot_filter, dùng logic mặc định")
    
    try:
        summary_op = operator_analysis_result.get("summary")
        
        # Lấy dữ liệu chi tiết
        details_ton_cu = operator_analysis_result.get("details_ton_cu", pd.DataFrame())
        details_ton_moi = operator_analysis_result.get("details_ton_moi", pd.DataFrame())
        details_roi = operator_analysis_result.get("details_roi_bai", pd.DataFrame())
        details_vao = operator_analysis_result.get("details_moi_vao", pd.DataFrame())
        
        # Gom biến động (Vào + Ra) thành 1 DataFrame
        df_bien_dong_list = []
        if not details_vao.empty:
            df_vao = details_vao.copy()
            df_vao['Loai_BD'] = 'VÀO'
            df_bien_dong_list.append(df_vao)
        if not details_roi.empty:
            df_roi = details_roi.copy()
            df_roi['Loai_BD'] = 'RA'
            df_bien_dong_list.append(df_roi)
        
        df_bien_dong = pd.concat(df_bien_dong_list, ignore_index=True) if df_bien_dong_list else pd.DataFrame()
        
        # Xác định danh sách hãng
        all_lines = set()
        if not details_ton_cu.empty and 'Lines' in details_ton_cu.columns:
            all_lines.update(details_ton_cu['Lines'].unique())
        if not details_ton_moi.empty and 'Lines' in details_ton_moi.columns:
            all_lines.update(details_ton_moi['Lines'].unique())
        
        for lines_name in sorted(all_lines):
            safe_name = lines_name.replace(' ', '_').replace('/', '_')
            file_path = hang_folder / f"TON_{safe_name}.xlsx"
            
            with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                # Lọc dữ liệu theo hãng
                df_cu_hang = details_ton_cu[details_ton_cu['Lines'] == lines_name].copy() if not details_ton_cu.empty and 'Lines' in details_ton_cu.columns else pd.DataFrame()
                df_moi_hang = details_ton_moi[details_ton_moi['Lines'] == lines_name].copy() if not details_ton_moi.empty and 'Lines' in details_ton_moi.columns else pd.DataFrame()
                df_bd_hang = df_bien_dong[df_bien_dong['Lines'] == lines_name].copy() if not df_bien_dong.empty and 'Lines' in df_bien_dong.columns else pd.DataFrame()
                
                time_config_note = ""
                if has_time_filter and Col.OPERATOR in df_cu_hang.columns:
                    # Lấy operator code đầu tiên trong Lines này
                    operators = df_cu_hang[Col.OPERATOR].unique().tolist() if not df_cu_hang.empty else []
                    if operators:
                        op_code = operators[0]
                        config = get_operator_time_config(op_code)
                        mode = config.get("mode", "full_day")
                        desc = config.get("description", "")
                        if mode == "split":
                            time_config_note = f"⚠️ Hãng này yêu cầu chia theo CA: {desc}"
                        else:
                            time_config_note = f"📅 Lấy nguyên ngày (8h-8h): {desc}"
                
                # Sheet 1: Tổng hợp của hãng này
                summary_data = {
                    'Hạng mục': ['Tồn Cũ', 'Tồn Mới', 'VÀO (Mới)', 'RA (Rời)', 'Biến Động'],
                    'Số lượng': [
                        len(df_cu_hang),
                        len(df_moi_hang),
                        len(df_bd_hang[df_bd_hang['Loai_BD'] == 'VÀO']) if not df_bd_hang.empty and 'Loai_BD' in df_bd_hang.columns else 0,
                        len(df_bd_hang[df_bd_hang['Loai_BD'] == 'RA']) if not df_bd_hang.empty and 'Loai_BD' in df_bd_hang.columns else 0,
                        len(df_moi_hang) - len(df_cu_hang)
                    ]
                }
                df_summary = pd.DataFrame(summary_data)
                
                # Thêm note về time config
                if time_config_note:
                    note_row = pd.DataFrame([{'Hạng mục': '', 'Số lượng': ''}, {'Hạng mục': time_config_note, 'Số lượng': ''}])
                    df_summary = pd.concat([df_summary, note_row], ignore_index=True)
                
                _write_sheet(writer, df_summary, "Tong_Hop", add_total=False)
                
                # Sheet 2: Tồn Cũ
                if not df_cu_hang.empty:
                    _write_sheet(writer, df_cu_hang, "Ton_Cu")
                
                # Sheet 3: Tồn Mới
                if not df_moi_hang.empty:
                    _write_sheet(writer, df_moi_hang, "Ton_Moi")
                
                # Sheet 4: Biến Động (VÀO + RA gom chung)
                if not df_bd_hang.empty:
                    _write_sheet(writer, df_bd_hang, "Bien_Dong")
            
            logging.info(f"[TON_THEO_HANG] Đã tạo file: {file_path.name}")
        
        # Tạo file tổng hợp TẤT CẢ HÃNG
        tong_hop_path = hang_folder / "0_TONG_HOP_TAT_CA.xlsx"
        with pd.ExcelWriter(tong_hop_path, engine='xlsxwriter') as writer:
            if isinstance(summary_op, pd.DataFrame) and not summary_op.empty:
                df_with_index = summary_op.reset_index()
                df_with_total = _add_total_row(df_with_index)
                df_with_total.to_excel(writer, sheet_name="Tong_Hop", index=False)
        
        logging.info(f"[TON_THEO_HANG] Đã tạo {len(all_lines)} file riêng theo hãng")
    except Exception as e:
        logging.error(f"Lỗi khi tạo TON_THEO_HANG: {e}")
    
    # ========================================
    # FILE 7_Email_Templates: Xuất theo template cho email hãng tàu (V5.2)
    # ========================================
    export_results = {}  # Store for optional email sending
    try:
        from reports.email_template_exporter import export_all_operators
        
        email_folder = report_folder / "7_Email_Templates"
        email_folder.mkdir(exist_ok=True)
        
        # Lấy dữ liệu biến động và tồn bãi
        df_bien_dong_email = pd.DataFrame()
        if inventory_change_results:
            df_vao = inventory_change_results.get('moi_vao_bai', pd.DataFrame())
            df_roi = inventory_change_results.get('da_roi_bai', pd.DataFrame())
            df_bien_dong_email = pd.concat([df_vao, df_roi], ignore_index=True) if not df_vao.empty or not df_roi.empty else pd.DataFrame()
        
        df_ton_bai_email = raw_data.get('ton_moi', pd.DataFrame())
        
        # Extract date string from report folder name
        folder_name = report_folder.name  # e.g., "Report_N12.01.2026_16h12"
        date_str = folder_name.replace("Report_", "").split("_")[0]  # "N12.01.2026"
        
        # Xuất files cho tất cả operators (với parallel processing)
        if not df_bien_dong_email.empty or not df_ton_bai_email.empty:
            df_ton_cu = raw_data.get('ton_cu', pd.DataFrame())
            df_gate_in = raw_data.get('gate_in', pd.DataFrame())
            df_gate_out = raw_data.get('gate_out', pd.DataFrame())
            
            if not df_bien_dong_email.empty:
                try:
                    from reports.email_template_exporter import enrich_with_raw_gate_data
                    df_bien_dong_email = enrich_with_raw_gate_data(
                        df_bien_dong_email, 
                        df_gate_in=df_gate_in, 
                        df_gate_out=df_gate_out
                    )
                    logging.info("[EMAIL_TEMPLATES] Đã bổ sung thông tin từ raw gate data")
                except Exception as enrich_err:
                    logging.warning(f"[EMAIL_TEMPLATES] Không thể enrich data: {enrich_err}")
            
            export_results = export_all_operators(
                df_bien_dong=df_bien_dong_email,
                df_ton_bai=df_ton_bai_email,
                date_str=date_str,
                output_dir=email_folder,
                parallel=True,
                df_ton_cu=df_ton_cu,
                df_gate_in=df_gate_in,
                df_gate_out=df_gate_out,
                enable_stacking_incoming_filter=True
            )
            logging.info(f"[EMAIL_TEMPLATES] Đã tạo files email cho {len(export_results)} hãng tại: {email_folder}")


            
            # Uncomment để bật tự động gửi email
            # try:
            #     from reports.email_sender import send_operator_emails
            #     from config import EMAIL_SETTINGS
            #     if EMAIL_SETTINGS.get("enabled", False):
            #         smtp_config = {
            #             'server': EMAIL_SETTINGS.get('smtp_server'),
            #             'port': EMAIL_SETTINGS.get('smtp_port'),
            #             'email': EMAIL_SETTINGS.get('sender_email'),
            #             'password': EMAIL_SETTINGS.get('sender_password')
            #         }
            #         send_results = send_operator_emails(export_results, date_str, smtp_config)
            #         logging.info(f"[AUTO_EMAIL] Đã gửi {sum(1 for r in send_results.values() if r.get('success'))} emails")
            # except Exception as email_err:
            #     logging.warning(f"[AUTO_EMAIL] Không thể gửi email tự động: {email_err}")
        else:
            logging.warning("[EMAIL_TEMPLATES] Không có dữ liệu để xuất email templates")
            
    except ImportError:
        logging.warning("[EMAIL_TEMPLATES] Module email_template_exporter chưa được cài đặt")
    except Exception as e:
        logging.error(f"Lỗi khi tạo EMAIL_TEMPLATES: {e}")
    
    # ========================================
    # FILE 8: MASTER_LOG.xlsx - Tra cứu
    # ========================================
    master_log_path = report_folder / "8. MASTER_LOG.xlsx"
    try:
        df_master = main_results.get('master_log', pd.DataFrame())
        if not df_master.empty:
            with pd.ExcelWriter(master_log_path, engine='xlsxwriter') as writer:
                _write_sheet(writer, df_master, "All_Moves")
                _write_sheet(writer, main_results.get('timeline'), "Timeline")
    except Exception as e:
        logging.error(f"Lỗi khi tạo MASTER_LOG: {e}")
    
    # ========================================
    # FILE 9: V5.1 MOVEMENT_SUMMARY.xlsx - Biến động chi tiết
    # ========================================
    if generate_full_movement_report is not None:
        movement_path = report_folder / "9. MOVEMENT_SUMMARY.xlsx"
        try:
            df_movement, balance_info = generate_full_movement_report(raw_data)
            with pd.ExcelWriter(movement_path, engine='xlsxwriter') as writer:
                # Sheet 1: Tổng hợp biến động
                _write_sheet(writer, df_movement, "Tong_Hop_Bien_Dong", add_total=False)
                
                # Sheet 2: Thông tin cân đối
                balance_df = pd.DataFrame([{
                    'Hạng mục': 'Tồn cũ',
                    'Giá trị': balance_info['ton_cu']
                }, {
                    'Hạng mục': 'Nhập (Gate In + Nhập tàu)',
                    'Giá trị': balance_info['nhap']
                }, {
                    'Hạng mục': 'Xuất (Gate Out + Xuất tàu)',
                    'Giá trị': balance_info['xuat']
                }, {
                    'Hạng mục': 'Tồn mới dự kiến',
                    'Giá trị': balance_info['ton_moi_du_kien']
                }, {
                    'Hạng mục': 'Tồn mới thực tế',
                    'Giá trị': balance_info['ton_moi_thuc_te']
                }, {
                    'Hạng mục': 'Chênh lệch',
                    'Giá trị': balance_info['chenh_lech']
                }, {
                    'Hạng mục': 'Trạng thái',
                    'Giá trị': '✅ CÂN ĐỐI' if balance_info['can_doi'] else '❌ CHƯA CÂN ĐỐI'
                }])
                _write_sheet(writer, balance_df, "Can_Doi", add_total=False)
                
                try:
                    from reports.movement_summary import create_vosco_movement_summary
                    df_vosco = create_vosco_movement_summary(raw_data, exclude_soc=True)
                    if not df_vosco.empty:
                        _write_sheet(writer, df_vosco, "Vosco_COC_Only", add_total=False)
                        logging.info("[MOVEMENT] Đã thêm sheet Vosco_COC_Only (loại trừ SOC)")
                except Exception as vosco_err:
                    logging.warning(f"[MOVEMENT] Không thể tạo sheet Vosco: {vosco_err}")
                
            logging.info(f"Đã tạo file Movement Summary: {movement_path}")
        except Exception as e:
            logging.error(f"Lỗi khi tạo MOVEMENT_SUMMARY: {e}")

    
    # ========================================
    # FILE 10: V5.1 ERRORS_V51.xlsx - Kiểm tra lỗi V5.1
    # ========================================
    v51_results = main_results.get('v51_checks', {})
    if v51_results:
        v51_path = report_folder / "10. ERRORS_V51.xlsx"
        try:
            with pd.ExcelWriter(v51_path, engine='xlsxwriter') as writer:
                has_data = False
                
                # Sheet: Container thiếu phương án
                df_missing_pa = v51_results.get('V51_missing_phuong_an', pd.DataFrame())
                if not df_missing_pa.empty:
                    _write_sheet(writer, df_missing_pa, "Thieu_Phuong_An")
                    has_data = True
                
                # Sheet: Container đổi OPR
                df_opr = v51_results.get('V51_opr_changes', pd.DataFrame())
                if not df_opr.empty:
                    _write_sheet(writer, df_opr, "Doi_OPR")
                    has_data = True
                
                # Sheet: Container đổi Size (NGHIÊM TRỌNG)
                df_size = v51_results.get('V51_size_changes', pd.DataFrame())
                if not df_size.empty:
                    _write_sheet(writer, df_size, "Doi_Size_NGHIEM_TRONG")
                    has_data = True
                
                # Sheet: Container đổi F/E
                df_fe = v51_results.get('V51_fe_changes', pd.DataFrame())
                if not df_fe.empty:
                    _write_sheet(writer, df_fe, "Doi_FE")
                    has_data = True
                
                if has_data:
                    logging.info(f"Đã tạo file Errors V5.1: {v51_path}")
                else:
                    logging.info("V5.1: Không có lỗi nào được phát hiện.")
        except Exception as e:
            logging.error(f"Lỗi khi tạo ERRORS_V51: {e}")
    
    logging.info(f"Hoàn tất! Báo cáo đã được tạo tại: {report_folder}")